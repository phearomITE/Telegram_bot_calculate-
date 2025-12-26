import io
from decimal import Decimal, ROUND_FLOOR, ROUND_HALF_UP


import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


from config import EXCHANGE_RATE_DEFAULT



USD_FORMAT = '"$"#,##0.00'
KHR_FORMAT = '"KHR "#,##0'
PERCENT_FORMAT = "0.00%"



# simple color mapping by sheet name (use ARGB hex)
SHEET_COLORS = {
    "Oil": "FFB18E00",              # yellow/brown
    "Powder Detergent": "FFFF00B8", # pink/magenta
    "Liquid Detergent": "FF00823B", # dark green
    "Milk": "FF3A6B24",             # green
    "Dishwash": "FFF4A57E",         # light orange
    "Fabric Softener": "FF8B4513",  # brown
    "Eco Dishwash": "FFF4A57E",     # same light orange
    "Toilet": "FF808080",           # gray
    "Data": "FF4F4F4F",             # fallback
}



def _fmt(x, symbol="$"):
    return x



def round2(x):
    """
    Custom rounding to 2 decimal places:
    - Look at 3rd decimal place
    - If 6-9: round UP
    - If 1-5: round DOWN
    - If 0 (or already 2 decimals): no change
    """
    if x is None:
        return None


    d = Decimal(str(x))
    scaled = (d * 1000).quantize(Decimal("1"), rounding=ROUND_FLOOR)
    third_digit = int(scaled % 10)


    if third_digit == 0:
        d2 = d.quantize(Decimal("0.01"), rounding=ROUND_FLOOR)
    elif third_digit >= 6:
        d2 = d.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    else:
        d2 = d.quantize(Decimal("0.01"), rounding=ROUND_FLOOR)


    return float(d2)



def round_weight(x):
    """
    Custom rounding for Weight per Ctn (whole numbers):
    - Look at first digit after decimal point
    - If 6-9: round UP to next integer
    - If 1-5: round DOWN to current integer
    - If .0 or no decimal: return as-is
    """
    if x is None:
        return None


    d = Decimal(str(x))
    integer_part = d.to_integral_value()
    fractional = d - integer_part


    if fractional == 0:
        return float(d)


    first_digit = int((fractional * 10).to_integral_value())


    if first_digit >= 6:
        return float(integer_part + 1)
    else:
        return float(integer_part)



def _to_float_money(x):
    if x is None:
        return None
    if isinstance(x, str):
        x = x.replace("$", "").strip()
        x = x.replace(",", "")
    return float(x)



def size_is_gram(data: dict) -> bool:
    raw = (data.get("size_raw") or "").lower()
    return "g" in raw and "ml" not in raw



def size_is_ml(data: dict) -> bool:
    raw = (data.get("size_raw") or "").lower()
    return "ml" in raw



def choose_sheet_name(data: dict) -> str:
    category = (data.get("category") or "").strip().lower()
    sub_cat = (data.get("sub_category") or "").strip().lower()


    oil_keywords = {
        "cooking oil",
        "oil",
        "palm oil",
        "vegetable oil",
        "coconut oil",
        "sunflower oil",
    }
    if category in oil_keywords:
        return "Oil"


    if (
        category in {"detergent", "powder detergent", "washing powder"}
        and sub_cat in {"powder", "powdered", "Powder", ""}
    ):
        return "Powder Detergent"


    if (
        category in {"detergent", "liquid detergent", "laundry liquid"}
        and sub_cat in {"liquid", "Liquid", ""}
    ):
        return "Liquid Detergent"


    milk_keywords = {
        "milk",
        "dairy milk",
        "fresh milk",
        "evaporated milk",
        "condensed milk",
        "Milk",
    }
    if category in milk_keywords:
        return "Milk"


    dishwash_keywords = {
        "dishwash",
        "dish wash",
        "dishwashing liquid",
        "Dishwash",
        "dishwashing",
    }
    if category in dishwash_keywords:
        return "Dishwash"


    fabric_keywords = {
        "fabric softener",
        "Fabric softener",
        "softener",
        "Fabric Softener",
        "fabric softner",
    }
    if category in fabric_keywords:
        return "Fabric Softener"


    eco_keywords = {
        "eco dishwash",
        "eco dishwashing",
        "eco-dishwash",
        "Eco dishwash",
        "Eco Dishwash",
    }
    if category in eco_keywords:
        return "Eco Dishwash"


    toilet_keywords = {
        "toilet",
        "toilet cleaner",
        "toilet bowl cleaner",
        "wc cleaner",
        "toilet liquid",
        "Toilet",
    }
    if category in toilet_keywords:
        return "Toilet"


    return "Data"



def calculate_fields(data: dict) -> dict:
    """
    Apply base rounding rules with round2 for monetary values.
    Excel will still do the formulas, but inputs are consistently rounded.
    """
    buy_in = _to_float_money(data.get("buy_in"))
    direct_disc_pct_input = data.get("direct_disc_pct")
    mark_up = _to_float_money(data.get("mark_up")) or 0
    size_val = data.get("size_ml") or 0
    price_unit_khr_input = data.get("price_unit_khr")


    exchange_rate = EXCHANGE_RATE_DEFAULT


    if buy_in is None:
        raise ValueError("Buy-in is required")
    if price_unit_khr_input is None:
        raise ValueError("Price Unit (KHR) is required")


    direct_disc_decimal = 0.0
    if direct_disc_pct_input is not None:
        direct_disc_decimal = round2(direct_disc_pct_input) / 100.0


    result = dict(data)
    result.update(
        size_ml=int(round(size_val)) if size_val is not None else None,
        buy_in=round2(buy_in),
        direct_disc_pct=direct_disc_decimal,
        mark_up=round2(mark_up),
        exchange_rate=float(exchange_rate),
        price_unit_khr=int(round(price_unit_khr_input)),
    )
    return result



def _row_from_data(data: dict) -> dict:
    """
    Build a row dict for DataFrame/Excel.
    Weight per Ctn will be calculated as formula in Excel.
    Other monetary fields are calculated in Excel using formulas with ROUND().
    """
    if size_is_gram(data):
        price_header = "Price / 100g"
    elif size_is_ml(data):
        price_header = "Price / 100ml"
    else:
        price_header = "Price / 100 unit"


    size_number = data.get("size_ml")
    packs_val = data.get("packs")


    buy_in = data.get("buy_in")
    mark_up = data.get("mark_up")


    row = {
        "Date": data.get("date"),
        "Address": data.get("address"),
        "Category": data.get("category"),
        "Sub-Category": data.get("sub_category"),
        "Brand": data.get("brand"),
        "Packaging": data.get("packaging"),
        "Size": size_number,
        "Packs": packs_val,
        "Weight per Ctn": None,  # Will be formula in Excel
        "Buy-in": buy_in,
        "Scheme(base)": data.get("scheme_base"),
        "FOC": data.get("foc"),
        "Discount(%)": None,
        "Discount($)": None,
        "Direct Disc.(%)": data.get("direct_disc_pct"),
        "Direct Disc($)": None,
        "Net Buy-in": None,
        "Price / 100 unit": None,
        "PriceHeader": price_header,
        "Mark - up": mark_up,
        "Sell Out ($)": None,
        "Exchange Rate": data.get("exchange_rate"),
        "Sell Out (KHR)": None,
        "Price Unit (KHR)": data.get("price_unit_khr"),
        "Margin/Unit (KHR)": None,
        "Price Ctn (KHR)": None,
        "Margin/Ctn (KHR)": None,
        "size_raw": data.get("size_raw"),
    }
    return row



def build_excel_from_sheet_dict(sheet_rows: dict) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)


    for sheet_name, rows in sheet_rows.items():
        if not rows:
            continue


        ws = wb.create_sheet(title=sheet_name)


        # sheet tab background color
        sheet_color = SHEET_COLORS.get(sheet_name, "FF4F4F4F")
        ws.sheet_properties.tabColor = sheet_color


        # section headers row 1
        section_headers = [
            ("A", "J", "PRODUCT INFO", sheet_color),
            ("K", "R", "WHOLESALE BUY-IN", sheet_color),
            ("S", "W", "WHOLESALE SELL-OUT", sheet_color),
            ("X", "AA", "RETAIL", sheet_color),
        ]
        for col_start, col_end, label, color in section_headers:
            ws.merge_cells(f"{col_start}1:{col_end}1")
            cell = ws[f"{col_start}1"]
            cell.value = label
            cell.font = Font(bold=True, size=11, color="FFFFFF")
            cell.fill = PatternFill(
                start_color=color, end_color=color, fill_type="solid"
            )
            cell.alignment = Alignment(horizontal="center", vertical="center")


        # Column headers row 2 with Id
        headers = [
            "Date",
            "Id",
            "Address",
            "Category",
            "Sub-Category",
            "Brand",
            "Packaging",
            "Size",
            "Packs",
            "Weight per Ctn",
            "Buy-in",
            "Scheme(base)",
            "FOC",
            "Discount(%)",
            "Discount($)",
            "Direct Disc.(%)",
            "Direct Disc($)",
            "Net Buy-in",
            "Price / 100 unit",
            "Mark - up",
            "Sell Out ($)",
            "Exchange Rate (KHR)",
            "Sell Out (KHR)",
            "Price Unit (KHR)",
            "Margin/Unit (KHR)",
            "Price Ctn (KHR)",
            "Margin/Ctn (KHR)",
        ]
        for col_num, header in enumerate(headers, start=1):
            cell = ws.cell(row=2, column=col_num)
            cell.value = header
            cell.font = Font(bold=True, size=10, color="FFFFFF")
            cell.fill = PatternFill(
                start_color="FF404040", end_color="FF404040", fill_type="solid"
            )
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
            cell.border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )


        df = pd.DataFrame(rows)
        df.columns = df.columns.str.strip()


        # sort by Date
        if "Date" in df.columns:
            df = df.sort_values(by=["Date"], ascending=True, na_position="last")


        weight_col_idx = headers.index("Weight per Ctn") + 1
        price_col_idx = headers.index("Price / 100 unit") + 1


        # Data rows
        for row_idx, (_, row_data) in enumerate(df.iterrows(), start=3):
            r = str(row_idx)


            if size_is_gram(row_data):
                ws.cell(row=2, column=weight_col_idx).value = "Weight per Ctn"
                ws.cell(row=2, column=price_col_idx).value = "Price / 100g"
            elif size_is_ml(row_data):
                ws.cell(row=2, column=weight_col_idx).value = "Weight per Ctn"
                ws.cell(row=2, column=price_col_idx).value = "Price / 100ml"
            else:
                ws.cell(row=2, column=weight_col_idx).value = "Weight per Ctn"
                ws.cell(row=2, column=price_col_idx).value = "Price / 100 unit"


            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=row_idx, column=col_idx)


                if header == "Date":
                    cell.value = row_data["Date"]


                elif header == "Id":
                    cell.value = row_idx - 2
                    cell.number_format = "0"


                elif header == "Size":
                    value = row_data["Size"]
                    cell.value = value
                    if size_is_ml(row_data):
                        cell.number_format = '#,##0" ml"'
                    elif size_is_gram(row_data):
                        cell.number_format = '#,##0" g"'
                    else:
                        cell.number_format = "#,##0"


                elif header == "Weight per Ctn":
                    # Size = H, Packs = I
                    if size_is_ml(row_data) or size_is_gram(row_data):
                        cell.value = (
                            f"=IF(H{r}=0,0,"
                            f"IF(MOD(ROUND((H{r}*I{r})/1000*10,0),10)>=6,"
                            f"ROUNDUP((H{r}*I{r})/1000,0),"
                            f"ROUNDDOWN((H{r}*I{r})/1000,0)))"
                        )
                        if size_is_ml(row_data):
                            cell.number_format = '#,##0" L"'
                        else:
                            cell.number_format = '#,##0" kg"'
                    else:
                        cell.value = (
                            f"=IF(H{r}=0,0,"
                            f"IF(MOD(ROUND(H{r}*I{r}*10,0),10)>=6,"
                            f"ROUNDUP(H{r}*I{r},0),"
                            f"ROUNDDOWN(H{r}*I{r},0)))"
                        )
                        cell.number_format = "#,##0"


                elif header == "Buy-in":
                    cell.value = float(row_data["Buy-in"])
                    cell.number_format = USD_FORMAT
                    cell.font = Font(color="FFED3F1C")


                elif header == "Discount(%)":
                    # Scheme(base)=L, FOC=M
                    cell.value = f"=IF((L{r}+M{r})=0,0,M{r}/(L{r}+M{r}))"
                    cell.number_format = PERCENT_FORMAT


                elif header == "Discount($)":
                    # Discount(%)=N, Buy-in=K
                    cell.value = f"=ROUND(N{r}*K{r},2)"
                    cell.number_format = USD_FORMAT


                elif header == "Direct Disc.(%)":
                    value = row_data["Direct Disc.(%)"]
                    cell.value = value
                    cell.number_format = PERCENT_FORMAT


                elif header == "Direct Disc($)":
                    # Direct Disc.(%)=P, Buy-in=K
                    cell.value = f"=ROUND(P{r}*K{r},2)"
                    cell.number_format = USD_FORMAT


                elif header == "Net Buy-in":
                    # Buy-in=K, Discount($)=O, Direct Disc($)=Q
                    cell.value = f"=ROUND(K{r}-(O{r}+Q{r}),2)"
                    cell.number_format = USD_FORMAT
                    cell.font = Font(color="FFED3F1C")


                elif header == "Price / 100 unit":
                    # Net Buy-in=R, Size=H, Packs=I
                    cell.value = (
                        f"=IF((H{r}*I{r})=0,0,"
                        f"ROUND(R{r}/((H{r}*I{r})/100),2))"
                    )
                    cell.number_format = USD_FORMAT


                elif header == "Mark - up":
                    value = row_data["Mark - up"]
                    cell.value = value
                    cell.number_format = USD_FORMAT


                elif header == "Sell Out ($)":
                    # Net Buy-in=R, Mark-up=T
                    cell.value = f"=ROUND(R{r}+T{r},2)"
                    cell.number_format = USD_FORMAT


                elif header == "Exchange Rate (KHR)":
                    value = row_data["Exchange Rate"]
                    cell.value = value
                    cell.number_format = KHR_FORMAT


                elif header == "Sell Out (KHR)":
                    # Sell Out ($)=U, Exchange Rate (KHR)=V
                    cell.value = f"=ROUND(U{r}*V{r},0)"
                    cell.number_format = KHR_FORMAT


                elif header == "Price Unit (KHR)":
                    value = row_data["Price Unit (KHR)"]
                    cell.value = value
                    cell.number_format = KHR_FORMAT


                elif header == "Margin/Unit (KHR)":
                    # Price Unit (KHR)=X, Sell Out (KHR)=W, Packs=I
                    cell.value = f"=ROUND(X{r}-(W{r}/I{r}),0)"
                    cell.number_format = KHR_FORMAT


                elif header == "Price Ctn (KHR)":
                    # Price Unit (KHR)=X, Packs=I
                    cell.value = f"=ROUND(X{r}*I{r},0)"
                    cell.number_format = KHR_FORMAT


                elif header == "Margin/Ctn (KHR)":
                    # Price Ctn (KHR)=Z, Sell Out (KHR)=W
                    cell.value = f"=ROUND(Z{r}-W{r},0)"
                    cell.number_format = KHR_FORMAT


                else:
                    value = row_data.get(header)
                    cell.value = value


                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.border = Border(
                    left=Side(style="thin"),
                    right=Side(style="thin"),
                    top=Side(style="thin"),
                    bottom=Side(style="thin"),
                )


        ws.row_dimensions[1].height = 25
        ws.row_dimensions[2].height = 30
        for col_num in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col_num)].width = 14


        # UPDATED: Freeze Panes to keep Buy-in column (K) visible
        # Freeze at L3: Keep rows 1-2 and columns A-K frozen
        ws.freeze_panes = "L3"


        ws.auto_filter.ref = f"A2:{get_column_letter(len(headers))}2"


    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
